"""
build_template.py — Parses Workout_Plan_Summer_24_.xlsx into a clean JSON template.
Forgets all old weights/times per user request. The template defines WHAT to do
on each day; the app collects actual performance and writes it to Supabase.
"""
import json
import openpyxl
from pathlib import Path

SRC = Path(__file__).parent.parent / "data" / "Workout_Plan_Summer_24_.xlsx"
OUT = Path(__file__).parent.parent / "app" / "template.json"

# Day blocks repeat every 14 rows. Within Week 1 columns (B-G = cols 2-7):
#   Row offset 0: "Week N"
#   Row offset 1: Day name (Monday..Friday)
#   Row offset 2: Muscle Group header w/ group name in col 5
#   Row offset 3: Exercise header (Sets/Reps/Set1/Set2/Set3  OR  Duration/Time for cardio)
#   Row offsets 4-12: exercises (variable count, NaN/blank ends block)

DAY_START_ROWS = [2, 16, 30, 44, 58]  # 1-indexed top row of each day's block
DAY_NAMES = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


def parse_day(ws, top_row):
    """Parse one day block starting at top_row. Returns dict with type & exercises."""
    muscle_group = ws.cell(row=top_row + 2, column=5).value
    header = ws.cell(row=top_row + 3, column=2).value  # "Exercise"
    col_b_header = ws.cell(row=top_row + 3, column=3).value  # "Sets" or "Duration (Miles)"

    is_cardio = "Duration" in str(col_b_header) if col_b_header else False
    day_type = "cardio" if is_cardio else "strength"

    exercises = []
    for r in range(top_row + 4, top_row + 13):
        name = ws.cell(row=r, column=2).value
        if not name or str(name).strip() == "":
            continue
        name = str(name).strip()
        if is_cardio:
            duration = ws.cell(row=r, column=3).value  # e.g. 1, 2, "3x"
            # Treadmill = run with miles target; everything else = timed core
            if "Treadmill" in name or "Run" in name:
                exercises.append({
                    "name": name,
                    "kind": "run",
                    "target_miles": float(duration) if isinstance(duration, (int, float)) else 1.0,
                })
            else:
                exercises.append({
                    "name": name,
                    "kind": "timed",
                    "sets": str(duration) if duration else "3x",
                    "target_seconds": 45,
                })
        else:
            sets = ws.cell(row=r, column=3).value
            reps = ws.cell(row=r, column=4).value
            exercises.append({
                "name": name,
                "kind": "lift",
                "sets": int(sets) if isinstance(sets, (int, float)) else 3,
                "reps": int(reps) if isinstance(reps, (int, float)) else 12,
            })

    return {
        "day_type": day_type,
        "muscle_group": str(muscle_group).strip() if muscle_group else "",
        "exercises": exercises,
    }


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    ws = wb.active

    days = {}
    for name, top in zip(DAY_NAMES, DAY_START_ROWS):
        days[name] = parse_day(ws, top)

    template = {
        "program_name": "Summer Strength + Conditioning",
        "program_weeks": 14,
        "schedule": days,
        "rest_days": ["Saturday", "Sunday"],
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(template, indent=2))
    print(f"Wrote {OUT}")
    print(json.dumps(template, indent=2))


if __name__ == "__main__":
    main()
